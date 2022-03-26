# -*- coding: utf-8 -*-
"""
Created on Fri Feb 25 13:25:05 2022

@author: Nisha Poudel
"""


import tkinter as tk
import openpyxl,xlrd
from openpyxl import Workbook
from tkinter.messagebox import showinfo
import pathlib


import datetime 
import time
from PIL import Image, ImageTk
#import facedetection
   
file= pathlib.Path("register.xlsx")
if file.exists():
    pass
else:
    file= Workbook()
    sheet= file.active
    sheet["A1"]= "ID"
    sheet["B1"]= "Name"
    sheet["C1"]= "Faculty"
    sheet["D1"]= "Batch"
    sheet["E1"]= "Email"
    file.save("register.xlsx")  
def save():
    
    ids= txtid.get()
    name= txtname.get()
    faculty= txtfaculty.get()
    batch= txtbatch.get()
    email= txtemail.get()
    
    print(ids)
    print(name)
    print(faculty)
    print(batch)
    print(email)
    
    file=openpyxl.load_workbook("register.xlsx")
    sheet = file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=ids)
    sheet.cell(column=2,row=sheet.max_row,value=name)
    sheet.cell(column=3,row=sheet.max_row,value=faculty)
    sheet.cell(column=4,row=sheet.max_row,value=batch)
    sheet.cell(column=5,row=sheet.max_row,value=email)
    file.save("register.xlsx")
    
    

    showinfo("Saved","Your Entry has been saved")
  
   

    
def registerme():
    window=tk.Tk()

    def tick():
        time_string = time.strftime('%H:%M:%S')
        clock.config(text=time_string)
        clock.after(200,tick)

    ts = time.time()
    date = datetime.datetime.fromtimestamp(ts).strftime('%d-%m-%Y')
    day,month,year=date.split("-")
    
    mont={'01':'January',
          '02':'February',
          '03':'March',
          '04':'April',
          '05':'May',
          '06':'June',
          '07':'July',
          '08':'August',
          '09':'September',
          '10':'October',
          '11':'November',
          '12':'December'
          }
    def takeimage_full():
        facedetection.TakeImage();
    #windows title
    window.title("Register.Fill the form")
    #windows stze
    window.geometry("600x600")
    window.resizable(False, False)
    
    """
    
    #frame 1 is for image ie header
    frame1=tk.Frame(window,width=200,height=100,bg="red")
    frame1.pack(fill=tk.BOTH)
    image=Image.open("1.jpg")
    resize_img=image.resize((500,130),Image.ANTIALIAS)
    photo=ImageTk.PhotoImage(resize_img)
    h1=tk.Label(frame1,image=photo)
    h1.place(x=0,y=0,width=500,height=100)
    
    #image2
    image2=Image.open("2.jpg")
    resize_img2=image2.resize((500,130),Image.ANTIALIAS)
    photo2=ImageTk.PhotoImage(resize_img2)
    h2=tk.Label(frame1,image=photo2)
    h2.place(x=500,y=0,width=500,height=100)
    """
    
    #frame 2 is for title
    frame2=tk.Frame(window,width=200,height=100)
    frame2.pack(fill=tk.BOTH)
    message1 = tk.Label(frame2, text="Face Recognition Based Attendance System" ,fg="green" ,width=80,height=2,font=('times', 20, ' bold '))
    #message1.place(x=80,y=10)
    message1.pack()
    
    #date 
    datef = tk.Label(frame2, text = day+"-"+mont[month]+"-"+year+"   ", fg="orange" ,width=80 ,height=1,font=('times', 20, ' bold '))
    #datef.place(x=100,y=160)
    datef.pack()
    #time
    clock = tk.Label(frame2,fg="orange" ,width=80 ,height=1,font=('times', 20, ' bold '))
    #clock.place(x=100,y=200)
    clock.pack()
    tick()
    
    #--------------------------------registration part start---------------------------------------------
    
    frame3=tk.Frame(window,width=200,height=100)
    frame3.pack(fill=tk.BOTH)
    message2 = tk.Label(frame3, text="Registration Form" ,fg="green" ,width=80,height=2,font=('times', 20, ' bold '))
    #message1.place(x=80,y=10)
    message2.pack()
    
    #resgister ko lagi  frame
    
    frame4=tk.Frame(window,width=200,height=300,bg="yellow")
    frame4.pack(fill=tk.BOTH)
    
    #tkinter variable for storing entries
    
    
    
    
    lbl1= tk.Label(frame4, text="Enter ID",width=8  ,height=1  ,fg="black",font=('times', 17, ' bold ') )
    lbl1.place(x=140, y=55)
  
    txtid = tk.Entry(frame4,width=20,fg="black",font=('times', 15, ' bold ' ))
    txtid.place(x=280, y=55)
    
    lbl2= tk.Label(frame4, text="Name",width=8 ,height=1  ,fg="black",font=('times', 17, ' bold ') )
    lbl2.place(x=140, y=90)
    
    txtname = tk.Entry(frame4,width=20,fg="black",font=('times', 15, ' bold '))
    txtname.place(x=280, y=90)
    
    lbl3= tk.Label(frame4, text="Faculty",width=8  ,height=1  ,fg="black",font=('times', 17, ' bold ') )
    lbl3.place(x=140, y=125)
    txtfaculty = tk.Entry(frame4,width=20,fg="black",font=('times', 15, ' bold '))
    txtfaculty.place(x=280, y=125)
    
    lbl4= tk.Label(frame4, text="Batch",width=8  ,height=1  ,fg="black",font=('times', 17, ' bold ') )
    lbl4.place(x=140, y=160)
    txtbatch = tk.Entry(frame4,width=20,fg="black",font=('times', 15, ' bold '))
    txtbatch.place(x=280, y=160)
    
    lbl5= tk.Label(frame4, text="Email",width=8  ,height=1  ,fg="black",font=('times', 17, ' bold ') )
    lbl5.place(x=140, y=195)
    txtemail = tk.Entry(frame4,width=20,fg="black",font=('times', 15, ' bold '))
    txtemail.place(x=280, y=195)
    
    button_save =tk.Button(frame4, text="Take Image",width="10",height="2", fg="red",command=takeimage_full)
    button_save.place(x=200,y=250)
    
    button_save =tk.Button(frame4, text="Save Profile", width="10",height="2",fg="red",command= save)
    button_save.place(x=300,y=250)
    
    
    
    
    window.mainloop();
    
    
    
    
        
        